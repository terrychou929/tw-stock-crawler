import os
import sys
from crawler import StockCrawler
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from openpyxl import Workbook
from utils import write_raw_data, style_summary_sheet

def main():
    # Check command line arguments
    if len(sys.argv) != 2:
        print("Please provide a stock code, e.g., python main.py 2330")
        sys.exit(1)
    
    stock_code = sys.argv[1]
    
    # Use multi-threading to fetch data
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Execute crawler tasks in parallel
        revenue_data = executor.submit(StockCrawler(stock_code).get_revenue)
        profit_ratio_data = executor.submit(StockCrawler(stock_code).get_profit_ratio)
        pe_ratio_data = executor.submit(StockCrawler(stock_code).get_pe_ratio)
        current_stock_price_data = executor.submit(StockCrawler(stock_code).get_current_stock_price)
        share_number_data = executor.submit(StockCrawler(stock_code).get_share_number)
        
        # Wait for all tasks to complete and retrieve results
        revenue = revenue_data.result()
        profit_ratio = profit_ratio_data.result()
        pe_ratio = pe_ratio_data.result()
        current_stock_price = current_stock_price_data.result()
        share_number = share_number_data.result()

    # Create output directory if it doesn't exist
    output_dir = "output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{output_dir}/{stock_code}_stock_data_{timestamp}.xlsx"
    wb = Workbook()
    
    # Remove default empty sheet
    wb.remove(wb.active)
    
    # Write raw data to sheets
    ws_revenue = wb.create_sheet("Revenue")
    write_raw_data(ws_revenue, revenue)
    
    ws_profit = wb.create_sheet("Profit Margin")
    write_raw_data(ws_profit, profit_ratio)
    
    ws_pe = wb.create_sheet("PE Ratio")
    write_raw_data(ws_pe, pe_ratio)
    
    ws_price = wb.create_sheet("Current Price")
    write_raw_data(ws_price, current_stock_price)
    
    ws_share = wb.create_sheet("Share")
    write_raw_data(ws_share, share_number)
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)  # Place it as the first sheet
    
    # Calculate required data
    current_price = current_stock_price["Price"].iloc[0] if not current_stock_price.empty else None
    share = share_number["Share"].iloc[0] if not share_number.empty else None
    latest_pe = pe_ratio["P/E Ratio"].iloc[0] if not pe_ratio.empty else None
    pe_min = pe_ratio["P/E Ratio"].min()
    pe_avg = pe_ratio["P/E Ratio"].mean()
    pe_max = pe_ratio["P/E Ratio"].max()
    profit_min = profit_ratio["Net Profit Margin"].min()
    profit_avg = profit_ratio["Net Profit Margin"].mean()
    profit_max = profit_ratio["Net Profit Margin"].max()
    
    # Calculate the sum of the latest 12 months' revenue
    if not revenue.empty and len(revenue) >= 12:
        latest_12_months_revenue = revenue["Revenue"].tail(12).sum()
    else:
        latest_12_months_revenue = None
        print("Warning: Insufficient revenue data for the last 12 months")
    
    # Write summary data
    summary_data = [
        ["Stock Code", stock_code],
        ["Current Price", current_price],
        ["Share Number", share],
        ["Latest P/E Ratio", latest_pe],
        ["", ""],  # Empty row
        ["Past 180 Weeks P/E Ratio", ""],
        ["Minimum", pe_min],
        ["Average", pe_avg],
        ["Maximum", pe_max],
        ["", ""],  # Empty row
        ["Past 5 Years Net Profit Margin", ""],
        ["Minimum", profit_min],
        ["Average", profit_avg],
        ["Maximum", profit_max],
        ["", ""],  # Empty row
        ["Price Prediction (Latest 12 Months Revenue Sum * 10 * Profit Margin / 2593 * P/E)", ""],
    ]
    
    # Calculate 9 predicted prices using the new formula
    if latest_12_months_revenue and all([pe_min, pe_avg, pe_max, profit_min, profit_avg, profit_max]):
        profit_ratios = [profit_min, profit_avg, profit_max]
        pe_ratios = [pe_min, pe_avg, pe_max]
        price_predictions = []
        for profit in profit_ratios:
            for pe in pe_ratios:
                predicted_price = (latest_12_months_revenue * 10 * (profit/100) / share) * pe
                price_predictions.append(predicted_price)
        
        prediction_labels = [
            "Min Profit * Min PE", "Min Profit * Avg PE", "Min Profit * Max PE",
            "Avg Profit * Min PE", "Avg Profit * Avg PE", "Avg Profit * Max PE",
            "Max Profit * Min PE", "Max Profit * Avg PE", "Max Profit * Max PE"
        ]
        for label, price in zip(prediction_labels, price_predictions):
            summary_data.append([label, f"{price:.2f}"])
    
    # Write data to Summary sheet
    for row in summary_data:
        ws_summary.append(row)
    
    # Apply styling to Summary sheet
    style_summary_sheet(ws_summary)
    
    # Save the file
    wb.save(output_file)
    print(f"Data saved to {output_file}")

if __name__ == "__main__":
    main()