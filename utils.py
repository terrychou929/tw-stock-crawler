from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def write_raw_data(ws, df):
    """
    Write DataFrame to a worksheet and auto-adjust column width
    """
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 1)
        ws.column_dimensions[column].width = adjusted_width

def style_summary_sheet(ws):
    """
    Apply styling to the Summary sheet
    """
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    subheader_fill = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                    top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, color="FFFFFF")
    subheader_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    
    for row in ws.rows:
        for cell in row:
            cell.border = border
            if cell.row == 1 or cell.row in [6, 11, 16]:  # Header rows
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            elif cell.row in [2, 3, 4, 7, 8, 9, 12, 13, 14, 17, 18, 19, 20, 21, 22 ,23, 24, 25]:  # Subheader or data rows
                if cell.column == 1:
                    cell.fill = subheader_fill
                    cell.font = subheader_font
                cell.alignment = align_center
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 1)
        ws.column_dimensions[column].width = adjusted_width